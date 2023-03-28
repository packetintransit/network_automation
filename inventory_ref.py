import getpass
from nornir import InitNornir
from nornir_netmiko.tasks import netmiko_send_command
from nornir_utils.plugins.functions import print_result
import openpyxl
from openpyxl.styles import Font

def initialize_nornir():
    nr = InitNornir(config_file="config.yaml")
    username = input("Please enter domain username: ")
    nr.inventory.defaults.username = username
    password = getpass.getpass()
    nr.inventory.defaults.password = password
    hotel_code = input("Please enter hotel code: ")
    nr.inventory.defaults.data = {"hotel_code": hotel_code}
    return nr

def dev_info(task):
    r = task.run(netmiko_send_command, command_string="show version", use_genie=True)
    task.host["facts"] = r.result
    serial = task.host['facts']['version']['chassis_sn']
    hoster = task.host['facts']['version']['hostname']
    interno = str(task.host['facts']['version']['number_of_intfs'])  # Convert to string
    image = task.host['facts']['version']['system_image']
    im_type = task.host['facts']['version']['image_type']
    operating = task.host['facts']['version']['os']
    up = task.host['facts']['version']['uptime']
    ver = task.host['facts']['version']['version']
    created = task.host['facts']['version']['compiled_date']
    task.host["facts"]["csvdata"] = ('BSH', task.host.hostname, hoster, serial,
                                     interno, image, im_type, operating, up, ver, created)

def main():
    nr = initialize_nornir()

    # Write Excel headers
    headers = ['Hotel Code', 'Management IP Address', 'Hostname', 'Serial Number',
               'Number of Interfaces', 'System Image', 'Image Type',
               'Operating System', 'Uptime', 'Version', 'Compiled Date']

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the header row and apply formatting
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=14)

    # Apply Excel native data filter to the header row
    sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    # Run the tasks and collect inventory data
    targets = nr.filter(hotel_code=nr.inventory.defaults.data["hotel_code"])
    result = targets.run(task=dev_info)

    # Add the inventory data to the Excel sheet
    for idx, device in enumerate(result.items(), start=2):
        _, task_result = device
        if 'facts' in task_result.host and 'csvdata' in task_result.host['facts']:
            dev_data = task_result.host['facts']['csvdata']
            for col_num, data in enumerate(dev_data, start=1):
                cell = sheet.cell(row=idx, column=col_num)
                cell.value = data

    # Save the Excel workbook
    workbook.save('inventory.xlsx')
    print_result(result)


if __name__ == '__main__':
    main()
